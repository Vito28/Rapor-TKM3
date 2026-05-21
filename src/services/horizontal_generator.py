from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import ExcelLayoutConfig
from models import GeneratedWorkbookResult
from utils.excel_reader import extract_student_names


def copy_cell_style(src_cell, dst_cell) -> None:
    """Copy style antar-workbook tanpa memakai _style internal openpyxl."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)

    if src_cell.hyperlink:
        dst_cell._hyperlink = copy(src_cell.hyperlink)

    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)


def safe_merge_cells(
    ws: Worksheet,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> None:
    try:
        ws.merge_cells(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column,
        )
    except IndexError:
        # Recovery untuk kasus border/style merge corrupt dari workbook lama.
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                ws.cell(row=row, column=col).border = Border()

        ws.merge_cells(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column,
        )


def copy_left_rubric(src_ws: Worksheet, dst_ws: Worksheet, config: ExcelLayoutConfig) -> None:
    """Copy bagian kiri A:E dari Sheet1 indikator ke file output baru."""
    for row in range(1, src_ws.max_row + 1):
        for col in range(1, config.left_copy_until_col + 1):
            src_cell = src_ws.cell(row=row, column=col)
            dst_cell = dst_ws.cell(row=row, column=col)
            dst_cell.value = src_cell.value
            copy_cell_style(src_cell, dst_cell)

    for col in range(1, config.left_copy_until_col + 1):
        col_letter = get_column_letter(col)
        dst_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width

    for row in range(1, src_ws.max_row + 1):
        dst_ws.row_dimensions[row].height = src_ws.row_dimensions[row].height

    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.max_col <= config.left_copy_until_col:
            safe_merge_cells(
                dst_ws,
                start_row=merged_range.min_row,
                start_column=merged_range.min_col,
                end_row=merged_range.max_row,
                end_column=merged_range.max_col,
            )


def style_generated_area(
    ws: Worksheet,
    max_row: int,
    first_col: int,
    last_col: int,
    config: ExcelLayoutConfig,
) -> None:
    thin = Side(style="thin", color="000000")
    dotted = Side(style="dotted", color="808080")

    for row in range(1, max_row + 1):
        for col in range(first_col, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if row >= config.data_start_row:
                cell.border = Border(left=dotted, right=dotted, top=dotted, bottom=dotted)
            else:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    name_fill = PatternFill("solid", fgColor="D9EAF7")
    score_fill = PatternFill("solid", fgColor="F2F2F2")

    for col in range(first_col, last_col + 1):
        ws.cell(config.name_row, col).font = Font(bold=True)
        ws.cell(config.name_row, col).fill = name_fill
        ws.cell(config.semester_row, col).font = Font(bold=True)
        ws.cell(config.semester_row, col).fill = name_fill
        ws.cell(config.score_row, col).font = Font(bold=True, size=8)
        ws.cell(config.score_row, col).fill = score_fill
        ws.column_dimensions[get_column_letter(col)].width = 5


def generate_student_blocks(
    ws: Worksheet,
    names: list[str],
    max_row: int,
    config: ExcelLayoutConfig,
) -> None:
    total_generated_cols = len(names) * config.columns_per_student
    last_col = config.first_student_col + total_generated_cols - 1

    style_generated_area(ws, max_row, config.first_student_col, last_col, config)

    for student_index, name in enumerate(names):
        start_col = config.first_student_col + student_index * config.columns_per_student
        end_col = start_col + config.columns_per_student - 1

        safe_merge_cells(
            ws,
            start_row=config.name_row,
            start_column=start_col,
            end_row=config.name_row,
            end_column=end_col,
        )
        ws.cell(config.name_row, start_col).value = name

        sem1_start = start_col
        sem1_end = start_col + 3
        sem2_start = start_col + 4
        sem2_end = start_col + 7

        safe_merge_cells(
            ws,
            start_row=config.semester_row,
            start_column=sem1_start,
            end_row=config.semester_row,
            end_column=sem1_end,
        )
        ws.cell(config.semester_row, sem1_start).value = config.semester_labels[0]

        safe_merge_cells(
            ws,
            start_row=config.semester_row,
            start_column=sem2_start,
            end_row=config.semester_row,
            end_column=sem2_end,
        )
        ws.cell(config.semester_row, sem2_start).value = config.semester_labels[1]

        for offset, label in enumerate(config.score_labels):
            ws.cell(config.score_row, sem1_start + offset).value = label
            ws.cell(config.score_row, sem2_start + offset).value = label

    ws.freeze_panes = f"{get_column_letter(config.first_student_col)}{config.data_start_row}"


def generate_horizontal_format(
    rubric_path: Path,
    student_data_path: Path,
    output_dir: Path,
    config: ExcelLayoutConfig,
    max_students: int | None = None,
) -> GeneratedWorkbookResult:
    """
    Generate file indikator horizontal dari dua workbook terpisah:
    1. rubric_path       = RAPOR HIJAU_INDIKATOR.xlsx, pakai Sheet1.
    2. student_data_path = file data murid/nilai baru, pakai Sheet1 default.
    """
    if not rubric_path.exists():
        raise FileNotFoundError(f"File indikator tidak ditemukan: {rubric_path}")
    if not student_data_path.exists():
        raise FileNotFoundError(f"File data murid tidak ditemukan: {student_data_path}")

    names = extract_student_names(
        workbook_path=student_data_path,
        sheet_name=config.student_data_sheet,
        name_header=config.name_header,
        max_students=max_students,
    )
    if not names:
        raise ValueError("Tidak ada nama murid yang bisa dibaca dari file data murid/nilai.")

    source_wb = load_workbook(rubric_path)
    if config.template_sheet not in source_wb.sheetnames:
        raise ValueError(
            f"Sheet template '{config.template_sheet}' tidak ditemukan. "
            f"Sheet tersedia: {source_wb.sheetnames}"
        )

    src_ws = source_wb[config.template_sheet]

    output_wb = Workbook()
    out_ws = output_wb.active
    out_ws.title = "Sheet1"

    copy_left_rubric(src_ws, out_ws, config)
    generate_student_blocks(out_ws, names, src_ws.max_row, config)

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"FORMAT_INDIKATOR_{len(names)}_MURID_{timestamp}.xlsx"
    output_wb.save(output_path)

    return GeneratedWorkbookResult(
        path=output_path,
        student_count=len(names),
        message=f"Format indikator berhasil dibuat untuk {len(names)} murid.",
    )
