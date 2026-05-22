from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import ExcelLayoutConfig


@dataclass(frozen=True)
class GenerateResult:
    path: Path
    message: str


def _style_header_cell(cell) -> None:
    side = Side(style="thin", color="999999")
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAD3")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def _style_body_cell(cell) -> None:
    side = Side(style="thin", color="DDDDDD")
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def _column_width(header: str, config: ExcelLayoutConfig) -> int:
    wide_headers = {
        "Nama peserta didik",
        "Tempat dan tanggal lahir",
        "Alamat",
        "Nama ayah",
        "Nama ibu",
        "Nama wali",
        "Alamat wali",
        "Tanggal pembagian rapor",
        "Tanggal masuk",
    }

    medium_headers = {
        "Nomor induk_NISN",
        "Diterima di kelompok",
        "Diterima tanggal",
        "Pekerjaan ayah",
        "Pekerjaan ibu",
        "Pekerjaan wali",
        "Nama panggilan",
    }

    if header in wide_headers:
        return 30

    if header in medium_headers:
        return 22

    if header in config.score_headers:
        return 14

    if header in {"S", "I", "A"}:
        return 8

    return 16


def _dummy_value(header: str, row_number: int, config: ExcelLayoutConfig):
    student_number = row_number - 1

    dummy_map = {
        "Nama peserta didik": f"Murid {student_number:02d}",
        "Nama panggilan": f"Murid{student_number}",
        "Nomor induk_NISN": f"2025{student_number:04d}",
        "Jenis kelamin": "L",
        "Tempat dan tanggal lahir": "Medan, 1 Januari 2020",
        "Agama": "Kristen",
        "Anak ke": 1,
        "Alamat": "Medan",
        "Telepon": "-",
        "Diterima di kelompok": "B4",
        "Diterima tanggal": "1 Juli 2025",
        "Nama ayah": f"Ayah Murid {student_number:02d}",
        "Nama ibu": f"Ibu Murid {student_number:02d}",
        "Pekerjaan ayah": "-",
        "Pekerjaan ibu": "-",
        "Nama wali": "-",
        "Alamat wali": "-",
        "Telepon wali": "-",
        "Pekerjaan wali": "-",
        "Kelompok": "B4",
        "Semester": config.semester_value,
        "T.P.": "2025/2026",
        "Kelakuan": "A",
        "S": "-",
        "I": "-",
        "A": "-",
        "Tanggal pembagian rapor": "20 Desember 2025",
        "Guru": "-",
        "Naik ke ": "-",
        "Tanggal masuk": "1 Juli 2025",
    }

    if header in dummy_map:
        return dummy_map[header]

    if header in config.score_headers:
        return round(8 + ((student_number % 7) * 0.2), 1)

    return ""


def generate_student_data_template(
    output_dir: Path,
    config: ExcelLayoutConfig,
    row_count: int,
    with_dummy_data: bool = False,
) -> GenerateResult:
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = config.student_data_sheet

    headers = config.student_data_headers

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _style_header_cell(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = _column_width(header, config)

    for row_idx in range(2, row_count + 2):
        for col_idx, header in enumerate(headers, start=1):
            value = _dummy_value(header, row_idx, config) if with_dummy_data else ""

            if not with_dummy_data:
                if header == "Semester":
                    value = config.semester_value
                elif header == "T.P.":
                    value = "2025/2026"
                elif header == "Kelompok":
                    value = "B4"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _style_body_cell(cell)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"TEMPLATE_DATA_MURID_NILAI_{row_count}_MURID_{timestamp}.xlsx"

    wb.save(output_path)

    return GenerateResult(
        path=output_path,
        message=f"Template data murid/nilai berhasil dibuat untuk {row_count} murid.",
    )