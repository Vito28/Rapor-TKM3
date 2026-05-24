from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any
import re

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.config import ExcelLayoutConfig
from src.models import GeneratedWorkbookResult
from src.utils.excel_reader import (
    normalize,
    normalize_header,
    nickname_from_name,
    read_rows_by_header,
)


# ==========================================================
# HEADER / WORKSHEET HELPERS
# ==========================================================

def header_map(ws: Worksheet) -> dict[str, int]:
    mapping: dict[str, int] = {}

    for col in range(1, ws.max_column + 1):
        header = normalize(ws.cell(1, col).value)

        if header:
            mapping[normalize_header(header)] = col

    return mapping


def get_col(headers: dict[str, int], header_name: str) -> int | None:
    return headers.get(normalize_header(header_name))


def write_if_header_exists(
    ws: Worksheet,
    headers: dict[str, int],
    row_idx: int,
    header_name: str,
    value: Any,
) -> bool:
    col = get_col(headers, header_name)

    if not col:
        return False

    ws.cell(row=row_idx, column=col).value = value
    return True


def write_to_first_existing_header(
    ws: Worksheet,
    headers: dict[str, int],
    row_idx: int,
    possible_headers: tuple[str, ...],
    value: Any,
) -> bool:
    for header_name in possible_headers:
        if write_if_header_exists(ws, headers, row_idx, header_name, value):
            return True

    return False


def clear_existing_data(ws: Worksheet) -> None:
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)


# ==========================================================
# STYLE HELPERS
# ==========================================================

def capture_row_styles(ws: Worksheet, source_row: int) -> list[dict[str, Any]]:
    styles: list[dict[str, Any]] = []

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(source_row, col)

        styles.append(
            {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
            }
        )

    return styles


def apply_row_styles(ws: Worksheet, row: int, styles: list[dict[str, Any]]) -> None:
    for col, style in enumerate(styles, start=1):
        cell = ws.cell(row, col)

        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.number_format = style["number_format"]
        cell.protection = copy(style["protection"])


# ==========================================================
# NORMALIZATION / ALIAS HELPERS
# ==========================================================

def canonical_header(header: str, config: ExcelLayoutConfig) -> str:
    cleaned = normalize(header)

    if not cleaned:
        return ""

    for alias, target in config.header_aliases.items():
        if normalize_header(alias) == normalize_header(cleaned):
            return target

    return cleaned


def build_source_row_with_aliases(
    source_row: dict[str, Any],
    config: ExcelLayoutConfig,
) -> dict[str, Any]:
    result: dict[str, Any] = {}

    for header, value in source_row.items():
        canonical = canonical_header(header, config)

        if canonical:
            result[canonical] = value

    return result


def _score_headers(config: ExcelLayoutConfig) -> set[str]:
    return {score_header for score_header, _predicate_header in config.score_to_predicate_headers}


def _predicate_headers(config: ExcelLayoutConfig) -> set[str]:
    return {predicate_header for _score_header, predicate_header in config.score_to_predicate_headers}


def merge_default_values(
    source_row: dict[str, Any],
    default_values: dict[str, object],
    config: ExcelLayoutConfig,
) -> dict[str, Any]:
    """
    Default dari UI hanya mengisi kalau data Excel kosong.

    Penting:
    - Nilai angka tidak boleh diisi dari default UI.
    - P1-P17 tidak boleh diisi dari default UI.
    - Kalau nilai input kosong, output database tetap kosong.
    """
    result = dict(source_row)
    normalized_defaults = build_source_row_with_aliases(default_values, config)

    blocked_headers = _score_headers(config) | _predicate_headers(config)

    for header, value in normalized_defaults.items():
        if header in blocked_headers:
            continue

        if value in ("", None):
            continue

        if result.get(header) in ("", None):
            result[header] = value

    return result


# ==========================================================
# SCORE / PREDICATE HELPERS
# ==========================================================

def normalize_number(value: Any) -> float | None:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = normalize(value)

    if not text or text == "-":
        return None

    # Format Indonesia:
    # 1.234,56 -> 1234.56
    # 8,5      -> 8.5
    if re.fullmatch(r"\d{1,3}(\.\d{3})+(,\d+)?", text):
        text = text.replace(".", "")

    text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def format_number_for_excel(value: float | None) -> float | int | None:
    if value is None:
        return None

    rounded = round(value, 2)

    if rounded.is_integer():
        return int(rounded)

    return rounded


def predicate_from_score(value: Any, config: ExcelLayoutConfig) -> str:
    score = normalize_number(value)

    if score is None:
        return ""

    rule = config.predicate

    if score < rule.kurang_below:
        return "D"

    if score <= rule.cukup_max:
        return "C"

    if score <= rule.baik_max:
        return "B"

    if score <= rule.sangat_baik_max:
        return "A"

    return ""


def compute_score_stats(
    row: dict[str, Any],
    config: ExcelLayoutConfig,
) -> tuple[float | None, float | None]:
    scores: list[float] = []

    for score_header, _predicate_header in config.score_to_predicate_headers:
        value = normalize_number(row.get(score_header))

        if value is not None:
            scores.append(value)

    if not scores:
        return None, None

    total = sum(scores)
    average = total / len(scores)

    return total, average


# ==========================================================
# INDICATOR NARRATIVE HELPERS
# ==========================================================

def normalize_narrative_mapping(
    manual_indicator_narratives_by_name: dict[str, dict[str, str]] | None,
) -> dict[str, dict[str, str]]:
    """
    UI boleh mengirim key nama dalam bentuk apa pun.
    Di sini semua key dinormalisasi supaya cocok dengan nama di data nilai.
    """
    if not manual_indicator_narratives_by_name:
        return {}

    normalized: dict[str, dict[str, str]] = {}

    for name, narratives in manual_indicator_narratives_by_name.items():
        key = normalize_header(name)

        if key:
            normalized[key] = narratives

    return normalized


def get_indicator_narratives_for_student(
    *,
    student_name: str,
    indicator_narratives_by_name: dict[str, dict[str, str]],
) -> dict[str, str]:
    if not student_name:
        return {}

    return indicator_narratives_by_name.get(normalize_header(student_name), {})


# ==========================================================
# WRITE ONE DATABASE ROW
# ==========================================================

def write_row_to_database(
    db_ws: Worksheet,
    headers: dict[str, int],
    output_row: int,
    index: int,
    source_row: dict[str, Any],
    config: ExcelLayoutConfig,
    default_values: dict[str, object] | None = None,
    indicator_narratives: dict[str, str] | None = None,
) -> None:
    default_values = default_values or {}
    indicator_narratives = indicator_narratives or {}

    normalized_source = build_source_row_with_aliases(source_row, config)

    normalized_source = merge_default_values(
        source_row=normalized_source,
        default_values=default_values,
        config=config,
    )

    name = normalize(normalized_source.get(config.name_header))
    nickname = normalize(normalized_source.get("Nama panggilan")) or nickname_from_name(name)

    # Kolom wajib/standar.
    write_if_header_exists(db_ws, headers, output_row, "No.", f"{index}.")
    write_if_header_exists(db_ws, headers, output_row, "Nama peserta didik", name)
    write_if_header_exists(db_ws, headers, output_row, "Nama panggilan", nickname)

    write_if_header_exists(
        db_ws,
        headers,
        output_row,
        "Semester",
        normalized_source.get("Semester") or config.semester_value,
    )

    # Copy semua kolom input/default yang memang ada di template database.
    # Kalau input kosong, hasil database tetap kosong karena sheet sudah dibersihkan.
    for source_header, value in normalized_source.items():
        write_if_header_exists(db_ws, headers, output_row, source_header, value)

    # Inject narasi indikator hasil submit UI.
    # Contoh header: Agama , Agama A, Jati Diri A, STEAM B, dst.
    for narrative_header, narrative_value in indicator_narratives.items():
        write_if_header_exists(db_ws, headers, output_row, narrative_header, narrative_value)

    # Auto predikat P1-P17 berdasarkan nilai angka.
    # Kalau nilai kosong / bukan angka, nilai dan predikat dibiarkan kosong.
    for score_header, predicate_header in config.score_to_predicate_headers:
        score_value = normalized_source.get(score_header)

        if score_value in (None, ""):
            continue

        numeric_score = normalize_number(score_value)

        if numeric_score is None:
            continue

        value_to_write = format_number_for_excel(numeric_score)
        predicate = predicate_from_score(numeric_score, config)

        write_if_header_exists(db_ws, headers, output_row, score_header, value_to_write)
        write_if_header_exists(db_ws, headers, output_row, predicate_header, predicate)

    # Auto total dan rata-rata hanya dari nilai yang ada.
    total, average = compute_score_stats(normalized_source, config)

    if total is not None:
        existing_total = (
            normalized_source.get("Jumlah semua")
            or normalized_source.get("Jumlah Semua")
            or normalized_source.get("Jumlah Nilai")
            or normalized_source.get("Jumlah_Nilai")
        )

        existing_average = (
            normalized_source.get("Rata-rata")
            or normalized_source.get("Rata Rata")
            or normalized_source.get("Rata_Rata")
        )

        if existing_total in (None, ""):
            write_to_first_existing_header(
                db_ws,
                headers,
                output_row,
                ("Jumlah semua", "Jumlah Semua", "Jumlah Nilai", "Jumlah_Nilai"),
                format_number_for_excel(total),
            )

        if existing_average in (None, ""):
            write_to_first_existing_header(
                db_ws,
                headers,
                output_row,
                ("Rata-rata", "Rata Rata", "Rata_Rata"),
                format_number_for_excel(average),
            )


# ==========================================================
# MAIN SERVICE
# ==========================================================

def generate_database(
    student_data_path: Path,
    database_template_path: Path,
    output_dir: Path,
    config: ExcelLayoutConfig,
    max_students: int,
    default_values: dict[str, object] | None = None,
    indicator_path: Path | None = None,
    target_semester_header: str = "SEMESTER I",
    manual_indicator_narratives_by_name: dict[str, dict[str, str]] | None = None,
) -> GeneratedWorkbookResult:
    """
    Generate database dari file data murid/nilai terpisah.

    Perilaku utama:
    - Data murid/nilai dibaca dari Sheet1 file baru, bukan Sheet3.
    - Nilai kosong tetap kosong.
    - P1-P17 dibuat otomatis hanya kalau nilai angka ada.
    - Narasi indikator dipakai dari hasil submit UI jika tersedia.
    - Jika nama indikator dan nama data nilai berbeda, UI harus mengirim mapping
      dalam manual_indicator_narratives_by_name.
    """
    default_values = default_values or {}

    # Parameter ini sengaja tetap ada untuk kompatibilitas dengan UI lama.
    # Narasi indikator yang dipakai sekarang berasal dari manual_indicator_narratives_by_name.
    _ = indicator_path
    _ = target_semester_header

    if not student_data_path.exists():
        raise FileNotFoundError(f"File data murid/nilai tidak ditemukan: {student_data_path}")

    if not database_template_path.exists():
        raise FileNotFoundError(f"Template database tidak ditemukan: {database_template_path}")

    rows = read_rows_by_header(
        workbook_path=student_data_path,
        sheet_name=config.student_data_sheet,
        required_header=config.name_header,
        max_rows=max_students,
    )

    if not rows:
        raise ValueError("Tidak ada data murid yang bisa dibaca dari file data murid/nilai.")

    indicator_narratives_by_name = normalize_narrative_mapping(
        manual_indicator_narratives_by_name
    )

    db_wb = load_workbook(database_template_path)
    db_ws = db_wb[config.database_sheet] if config.database_sheet in db_wb.sheetnames else db_wb.active

    headers = header_map(db_ws)

    if not headers:
        raise ValueError("Header database tidak terbaca. Pastikan baris 1 berisi nama kolom.")

    style_source_row = 2 if db_ws.max_row >= 2 else 1
    row_styles = capture_row_styles(db_ws, style_source_row)

    clear_existing_data(db_ws)

    for index, source_row in enumerate(rows, start=1):
        output_row = index + 1

        normalized_source = build_source_row_with_aliases(source_row, config)
        student_name = normalize(normalized_source.get(config.name_header))

        indicator_narratives = get_indicator_narratives_for_student(
            student_name=student_name,
            indicator_narratives_by_name=indicator_narratives_by_name,
        )

        apply_row_styles(db_ws, output_row, row_styles)

        write_row_to_database(
            db_ws=db_ws,
            headers=headers,
            output_row=output_row,
            index=index,
            source_row=source_row,
            config=config,
            default_values=default_values,
            indicator_narratives=indicator_narratives,
        )

    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"DATABASE_SEMESTER_1_{len(rows)}_MURID_{timestamp}.xlsx"

    db_wb.save(output_path)

    if indicator_narratives_by_name:
        indicator_message = " Narasi indikator diisi dari hasil submit UI."
    else:
        indicator_message = " Narasi indikator kosong karena belum ada hasil submit indikator dari UI."

    return GeneratedWorkbookResult(
        path=output_path,
        student_count=len(rows),
        message=(
            f"Database semester 1 berhasil dibuat untuk {len(rows)} murid. "
            f"Nilai dan P1-P17 sudah diproses otomatis."
            f"{indicator_message}"
        ),
    )