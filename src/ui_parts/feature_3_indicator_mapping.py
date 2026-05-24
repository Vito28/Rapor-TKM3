from __future__ import annotations

from datetime import datetime
from difflib import get_close_matches
from pathlib import Path
from typing import Any
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.services.runtime_files import read_bytes, save_uploaded_file
from src.utils.excel_reader import extract_student_names
from src.ui_parts.common import default_indicator_path


CHECK_SYMBOLS = {"✓", "✔", "☑", "√", "ü"}

LEVEL_COLUMNS = ("BB", "MB", "BSH", "BSB")
GOOD_LEVEL = "BSB"
BAD_LEVEL = "BB"
CAPABILITY_TEXT_COLUMN = 2

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SECTION_CONFIGS_UI = (
    {
        "marker": "I. ELEMEN NILAI AGAMA DAN BUDI PEKERTI",
        "name": "Agama",
        "summary_header": "Agama ",
        "category_prefix": "Agama",
        "summary_template": "Pada umumnya nilai agama dan moral {nickname} sudah baik.",
        "guidance_word": "nasihat",
    },
    {
        "marker": "II. ELEMEN JATI DIRI",
        "name": "Jati Diri",
        "summary_header": "Jati Diri ",
        "category_prefix": "Jati Diri",
        "summary_template": "Pada umumnya kemampuan jati diri {nickname} sudah baik.",
        "guidance_word": "latihan",
    },
    {
        "marker": "III. ELEMEN DASAR-DASAR LITERASI, MATEMATIKA, SAINS, TEKNOLOGI, REKAYASA DAN SENI",
        "name": "STEAM",
        "summary_header": "STEAM ",
        "category_prefix": "STEAM",
        "summary_template": (
            "Pada umumnya kemampuan dasar-dasar literasi, matematika, sains, "
            "teknologi, rekayasa dan seni {nickname} sudah baik."
        ),
        "guidance_word": "latihan",
    },
)


# ==========================================================
# TEXT HELPERS
# ==========================================================

def normalize_text(value) -> str:
    if value is None:
        return ""

    return " ".join(str(value).strip().split())


def normalize_key(value) -> str:
    return normalize_text(value).casefold()


def nickname_from_name(name: str) -> str:
    name = normalize_text(name)

    if not name:
        return "anak"

    return name.split()[0]


def safe_widget_key(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-zA-Z0-9_]+", "_", text)

    return text[:100]


def coerce_bool(value) -> bool:
    if isinstance(value, bool):
        return value

    if value is None:
        return False

    text = str(value).strip().lower()

    return text in {"true", "1", "yes", "y", "✓", "✔", "v"}


# ==========================================================
# EXCEL READ HELPERS
# ==========================================================

def get_merged_value(ws: Worksheet, row: int, col: int):
    cell = ws.cell(row=row, column=col)

    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value

    return None


def get_cell_text(ws: Worksheet, row: int, col: int) -> str:
    return normalize_text(get_merged_value(ws, row, col))


def is_checkmark(formula_value, real_value) -> bool:
    values = []

    if formula_value is not None:
        values.append(str(formula_value).strip())

    if real_value is not None:
        values.append(str(real_value).strip())

    for value in values:
        raw = value.strip()
        upper = raw.upper().replace(" ", "")

        if raw in CHECK_SYMBOLS:
            return True

        if "UNICHAR(10003)" in upper:
            return True

        if "UNICHAR(10004)" in upper:
            return True

        if "CHAR(252)" in upper:
            return True

    return False


def is_letter_label(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Z]\.?", normalize_text(value).upper()))


def is_number_label(value: str) -> bool:
    return bool(re.fullmatch(r"\d+\.", normalize_text(value)))


# ==========================================================
# NARRATIVE HELPERS
# ==========================================================

def clean_capability_text(text: str) -> str:
    text = normalize_text(text).rstrip(".")
    text = re.sub(r"^\d+\.\s*", "", text)

    if text:
        text = text[0].lower() + text[1:]

    return text


def join_phrases(items: list[str]) -> str:
    cleaned = []

    for item in items:
        text = clean_capability_text(item)

        if text:
            cleaned.append(text)

    if not cleaned:
        return ""

    if len(cleaned) == 1:
        return cleaned[0]

    if len(cleaned) == 2:
        return f"{cleaned[0]} dan {cleaned[1]}"

    return f"{', '.join(cleaned[:-1])}, dan {cleaned[-1]}"


def make_narrative(good_texts: list[str], bad_texts: list[str], guidance_word: str) -> str:
    good_part = join_phrases(good_texts)
    bad_part = join_phrases(bad_texts)

    if good_part and bad_part:
        return (
            f"Dalam hal {good_part} sangat baik, "
            f"namun dalam hal {bad_part} masih perlu bimbingan dan {guidance_word}."
        )

    if good_part:
        return f"Dalam hal {good_part} sudah sangat baik."

    if bad_part:
        return f"Dalam hal {bad_part} masih perlu bimbingan dan {guidance_word}."

    return ""


# ==========================================================
# INDICATOR STRUCTURE DETECTION
# ==========================================================

def find_semester_row(ws: Worksheet, target_semester_header: str, config: Any) -> int:
    if config.semester_row <= ws.max_row:
        for col in range(1, ws.max_column + 1):
            value = get_cell_text(ws, config.semester_row, col).upper()

            if value == target_semester_header.upper():
                return config.semester_row

    for row in range(1, min(ws.max_row, 40) + 1):
        for col in range(1, ws.max_column + 1):
            value = get_cell_text(ws, row, col).upper()

            if value == target_semester_header.upper():
                return row

    raise ValueError(f"Header {target_semester_header} tidak ditemukan di file indikator.")


def build_indicator_semester_block(
    ws: Worksheet,
    semester_row: int,
    start_col: int,
    end_col: int,
    target_semester_header: str,
    config: Any,
) -> dict | None:
    semester_name = get_cell_text(ws, semester_row, start_col).upper()

    if semester_name != target_semester_header.upper():
        return None

    score_columns = {}

    for col in range(start_col, end_col + 1):
        level = get_cell_text(ws, config.score_row, col).upper()

        if level in LEVEL_COLUMNS:
            score_columns[level] = col

    if not all(level in score_columns for level in LEVEL_COLUMNS):
        return None

    name = ""

    for col in range(start_col, end_col + 1):
        name = get_cell_text(ws, config.name_row, col)

        if name:
            break

    if not name:
        return None

    return {
        "name": name,
        "start_col": start_col,
        "end_col": end_col,
        "score_columns": score_columns,
    }


def find_indicator_semester_blocks(
    ws: Worksheet,
    target_semester_header: str,
    config: Any,
) -> list[dict]:
    semester_row = find_semester_row(ws, target_semester_header, config)

    blocks = []
    used_ranges = set()

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row != semester_row:
            continue

        value = get_cell_text(ws, merged_range.min_row, merged_range.min_col).upper()

        if value != target_semester_header.upper():
            continue

        start_col = merged_range.min_col
        end_col = merged_range.max_col
        used_ranges.add((start_col, end_col))

        block = build_indicator_semester_block(
            ws=ws,
            semester_row=semester_row,
            start_col=start_col,
            end_col=end_col,
            target_semester_header=target_semester_header,
            config=config,
        )

        if block:
            blocks.append(block)

    for col in range(1, ws.max_column + 1):
        value = get_cell_text(ws, semester_row, col).upper()

        if value != target_semester_header.upper():
            continue

        if any(start <= col <= end for start, end in used_ranges):
            continue

        block = build_indicator_semester_block(
            ws=ws,
            semester_row=semester_row,
            start_col=col,
            end_col=col + 3,
            target_semester_header=target_semester_header,
            config=config,
        )

        if block:
            blocks.append(block)

    blocks.sort(key=lambda item: item["start_col"])

    return blocks


def find_section_rows(ws: Worksheet) -> dict[str, tuple[int, int]]:
    markers = []

    for row in range(1, ws.max_row + 1):
        label = get_cell_text(ws, row, 1).upper()

        for section_config in SECTION_CONFIGS_UI:
            if label.startswith(section_config["marker"].upper()):
                markers.append((row, section_config))
                break

    if not markers:
        raise ValueError("Marker elemen I/II/III tidak ditemukan di file indikator.")

    result = {}

    for index, (start_row, section_config) in enumerate(markers):
        end_row = markers[index + 1][0] - 1 if index + 1 < len(markers) else ws.max_row
        result[section_config["name"]] = (start_row, end_row)

    return result


def detect_indicator_categories(
    ws: Worksheet,
    section_start: int,
    section_end: int,
) -> dict[str, list[int]]:
    categories = {}
    current_letter = ""

    for row in range(section_start + 1, section_end + 1):
        label = get_cell_text(ws, row, 1).upper().replace(" ", "")
        text = get_cell_text(ws, row, CAPABILITY_TEXT_COLUMN)

        if is_letter_label(label):
            current_letter = label.replace(".", "")
            categories[current_letter] = []
            continue

        if current_letter and is_number_label(label) and text:
            categories[current_letter].append(row)

    return categories


# ==========================================================
# INDICATOR PREVIEW DATAFRAME
# ==========================================================

def normalize_indicator_preview_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    required_columns = {"Nama indikator", "Elemen", "Kategori", "Indikator", *LEVEL_COLUMNS}
    missing = required_columns - set(df.columns)

    if missing:
        raise ValueError(f"File preview indikator tidak valid. Kolom hilang: {', '.join(sorted(missing))}")

    output = df.copy()

    if "__row_id" not in output.columns:
        output["__row_id"] = [f"row_{idx}" for idx in range(len(output))]

    if "No." not in output.columns:
        output["No."] = ""

    for level in LEVEL_COLUMNS:
        output[level] = output[level].map(coerce_bool)

    ordered_columns = [
        "__row_id",
        "Nama indikator",
        "Elemen",
        "Kategori",
        "No.",
        "Indikator",
        *LEVEL_COLUMNS,
    ]

    return output[ordered_columns]


def try_load_flat_indicator_preview(indicator_path: Path) -> pd.DataFrame | None:
    try:
        df = pd.read_excel(indicator_path)
    except Exception:
        return None

    required_columns = {"Nama indikator", "Elemen", "Kategori", "Indikator", *LEVEL_COLUMNS}

    if required_columns.issubset(set(df.columns)):
        return normalize_indicator_preview_dataframe(df)

    return None


def load_indicator_preview_dataframe(
    *,
    indicator_path: Path,
    config: Any,
    target_semester_header: str,
) -> pd.DataFrame:
    flat_preview = try_load_flat_indicator_preview(indicator_path)

    if flat_preview is not None:
        return flat_preview

    wb_formula = load_workbook(indicator_path, data_only=False)
    wb_value = load_workbook(indicator_path, data_only=True)

    if config.template_sheet not in wb_formula.sheetnames:
        raise ValueError(
            f"Sheet indikator '{config.template_sheet}' tidak ditemukan. "
            f"Sheet tersedia: {wb_formula.sheetnames}"
        )

    ws_formula = wb_formula[config.template_sheet]
    ws_value = wb_value[config.template_sheet]

    blocks = find_indicator_semester_blocks(
        ws=ws_formula,
        target_semester_header=target_semester_header,
        config=config,
    )

    if not blocks:
        raise ValueError(f"Tidak ada blok {target_semester_header} yang valid.")

    section_ranges = find_section_rows(ws_formula)

    rows = []
    row_number = 1

    for block in blocks:
        student_name = block["name"]
        score_columns = block["score_columns"]

        for section_config in SECTION_CONFIGS_UI:
            section_name = section_config["name"]
            start_row, end_row = section_ranges[section_name]
            categories = detect_indicator_categories(ws_formula, start_row, end_row)

            for category_letter, indicator_rows in categories.items():
                for indicator_row in indicator_rows:
                    indicator_text = get_cell_text(ws_formula, indicator_row, CAPABILITY_TEXT_COLUMN)
                    no_value = get_cell_text(ws_formula, indicator_row, 1)

                    if not indicator_text:
                        continue

                    row_data = {
                        "__row_id": f"{normalize_key(student_name)}_{row_number}",
                        "Nama indikator": student_name,
                        "Elemen": section_name,
                        "Kategori": category_letter,
                        "No.": no_value,
                        "Indikator": indicator_text,
                    }

                    for level in LEVEL_COLUMNS:
                        col = score_columns[level]
                        formula_value = get_merged_value(ws_formula, indicator_row, col)
                        real_value = get_merged_value(ws_value, indicator_row, col)
                        row_data[level] = is_checkmark(formula_value, real_value)

                    rows.append(row_data)
                    row_number += 1

    return normalize_indicator_preview_dataframe(pd.DataFrame(rows))


# ==========================================================
# MINIMAL FORM TABLE VALIDATION / UPDATE
# ==========================================================

def selected_levels_from_row(row) -> list[str]:
    return [level for level in LEVEL_COLUMNS if coerce_bool(row.get(level))]


def count_level_in_group(
    *,
    dataframe: pd.DataFrame,
    student_name: str,
    element: str,
    category: str,
    level: str,
    exclude_row_id: str | None = None,
) -> int:
    group_df = dataframe[
        (dataframe["Nama indikator"].map(normalize_key) == normalize_key(student_name))
        & (dataframe["Elemen"].astype(str) == str(element))
        & (dataframe["Kategori"].astype(str) == str(category))
    ]

    if exclude_row_id:
        group_df = group_df[group_df["__row_id"] != exclude_row_id]

    if group_df.empty:
        return 0

    return int(group_df[level].fillna(False).astype(bool).sum())


def validate_category_edit(
    *,
    current_category_df: pd.DataFrame,
    edited_category_df: pd.DataFrame,
) -> list[str]:
    errors: list[str] = []

    edited_by_id = {
        str(row["__row_id"]): row
        for _, row in edited_category_df.iterrows()
    }

    for _, current_row in current_category_df.iterrows():
        row_id = str(current_row.get("__row_id"))
        edited_row = edited_by_id.get(row_id)

        if edited_row is None:
            continue

        current_levels = selected_levels_from_row(current_row)
        proposed_levels = selected_levels_from_row(edited_row)

        row_no = normalize_text(current_row.get("No."))
        indicator = normalize_text(current_row.get("Indikator"))

        if len(proposed_levels) > 1:
            errors.append(
                f"No. {row_no}: satu indikator hanya boleh memilih satu dari BB/MB/BSH/BSB."
            )
            continue

        # Kalau sudah ada pilihan, user harus kosongkan dulu.
        # Jadi tidak boleh langsung pindah dari BB ke MB dalam sekali Apply.
        if len(current_levels) == 1 and len(proposed_levels) == 1:
            if current_levels[0] != proposed_levels[0]:
                errors.append(
                    f"No. {row_no}: pilihan aktif masih {current_levels[0]}. "
                    f"Kosongkan dulu sebelum mengganti ke {proposed_levels[0]}."
                )

        # Kalau data awal konflik lebih dari satu centang, wajib dikosongkan dulu.
        if len(current_levels) > 1 and len(proposed_levels) > 0:
            errors.append(
                f"No. {row_no}: baris masih konflik. Kosongkan semua pilihan dulu."
            )

    bb_count = int(edited_category_df["BB"].fillna(False).astype(bool).sum())
    bsb_count = int(edited_category_df["BSB"].fillna(False).astype(bool).sum())

    if bb_count > 1:
        errors.append("BB hanya boleh dicentang 1 kali dalam kategori ini.")

    if bsb_count > 1:
        errors.append("BSB hanya boleh dicentang 1 kali dalam kategori ini.")

    return errors


def apply_category_edit(
    *,
    edited_category_df: pd.DataFrame,
) -> None:
    if "indicator_editor_df" not in st.session_state:
        return

    master_df = st.session_state["indicator_editor_df"].copy()

    for _, edited_row in edited_category_df.iterrows():
        row_id = str(edited_row.get("__row_id", ""))

        if not row_id:
            continue

        mask = master_df["__row_id"] == row_id

        if not mask.any():
            continue

        for level in LEVEL_COLUMNS:
            master_df.loc[mask, level] = coerce_bool(edited_row.get(level))

    st.session_state["indicator_editor_df"] = master_df


def clear_category_edit(
    *,
    category_df: pd.DataFrame,
) -> None:
    if "indicator_editor_df" not in st.session_state:
        return

    master_df = st.session_state["indicator_editor_df"].copy()

    for _, row in category_df.iterrows():
        row_id = str(row.get("__row_id", ""))
        mask = master_df["__row_id"] == row_id

        for level in LEVEL_COLUMNS:
            master_df.loc[mask, level] = False

    st.session_state["indicator_editor_df"] = master_df
    st.success("Pilihan pada kategori ini berhasil dikosongkan.")


def render_compact_category_editor(
    *,
    selected_name: str,
    section_name: str,
    category_letter: str,
    category_df: pd.DataFrame,
) -> None:
    st.markdown(
        f"""
<div style="
    background:#F6F8FA;
    border:1px solid #E5E7EB;
    border-radius:8px;
    padding:8px 10px;
    margin-top:10px;
    margin-bottom:8px;
">
    <b>Kategori {category_letter}</b>
</div>
""",
        unsafe_allow_html=True,
    )

    bb_count = int(category_df["BB"].fillna(False).astype(bool).sum())
    bsb_count = int(category_df["BSB"].fillna(False).astype(bool).sum())

    note_parts = []

    if bb_count >= 1:
        note_parts.append("BB sudah ada 1 di kategori ini.")

    if bsb_count >= 1:
        note_parts.append("BSB sudah ada 1 di kategori ini.")

    if note_parts:
        st.markdown(
            "<br>".join(
                f"<span style='color:#B00020;font-size:13px;font-weight:700'>{note}</span>"
                for note in note_parts
            ),
            unsafe_allow_html=True,
        )

    editor_columns = [
        "__row_id",
        "No.",
        "Indikator",
        "BB",
        "MB",
        "BSH",
        "BSB",
    ]

    compact_df = category_df[editor_columns].copy()

    form_key = (
        f"form_indicator_"
        f"{safe_widget_key(selected_name)}_"
        f"{safe_widget_key(section_name)}_"
        f"{safe_widget_key(str(category_letter))}"
    )

    with st.form(form_key, clear_on_submit=False):
        edited_category_df = st.data_editor(
            compact_df,
            use_container_width=True,
            hide_index=True,
            height=min(360, 42 + (len(compact_df) * 36)),
            column_order=["No.", "Indikator", "BB", "MB", "BSH", "BSB"],
            column_config={
                "__row_id": None,
                "No.": st.column_config.Column(
                    "No.",
                    disabled=True,
                    width="small",
                ),
                "Indikator": st.column_config.Column(
                    "Indikator",
                    disabled=True,
                    width="large",
                ),
                "BB": st.column_config.CheckboxColumn(
                    "BB",
                    width="small",
                ),
                "MB": st.column_config.CheckboxColumn(
                    "MB",
                    width="small",
                ),
                "BSH": st.column_config.CheckboxColumn(
                    "BSH",
                    width="small",
                ),
                "BSB": st.column_config.CheckboxColumn(
                    "BSB",
                    width="small",
                ),
            },
            disabled=["__row_id", "No.", "Indikator"],
        )

        apply_col, clear_col = st.columns([1, 1])

        with apply_col:
            apply_clicked = st.form_submit_button(
                "Apply Kategori",
                use_container_width=True,
            )

        with clear_col:
            clear_clicked = st.form_submit_button(
                "Kosongkan Kategori",
                use_container_width=True,
            )

    if clear_clicked:
        clear_category_edit(category_df=category_df)

    if apply_clicked:
        validation_errors = validate_category_edit(
            current_category_df=category_df,
            edited_category_df=edited_category_df,
        )

        if validation_errors:
            st.error("Pilihan kategori belum valid.")
            for error in validation_errors:
                st.write(f"- {error}")
        else:
            apply_category_edit(
                edited_category_df=edited_category_df,
            )
            st.success("Pilihan kategori berhasil disimpan.")


def validate_indicator_editor(indicator_df: pd.DataFrame) -> list[str]:
    errors = []

    if indicator_df.empty:
        return ["Data indikator kosong."]

    for row_idx, row in indicator_df.iterrows():
        checked_count = sum(coerce_bool(row.get(level)) for level in LEVEL_COLUMNS)

        if checked_count > 1:
            errors.append(
                f"Baris {row_idx + 1}: satu indikator hanya boleh memiliki satu pilihan "
                f"BB/MB/BSH/BSB."
            )

    grouped = indicator_df.groupby(["Nama indikator", "Elemen", "Kategori"], dropna=False)

    for (student_name, element, category), group in grouped:
        bb_count = int(group["BB"].fillna(False).astype(bool).sum())
        bsb_count = int(group["BSB"].fillna(False).astype(bool).sum())

        if bb_count > 1:
            errors.append(
                f"{student_name} - {element} {category}: BB dicentang {bb_count} kali. Maksimal 1."
            )

        if bsb_count > 1:
            errors.append(
                f"{student_name} - {element} {category}: BSB dicentang {bsb_count} kali. Maksimal 1."
            )

    return errors


def save_indicator_editor_dataframe(
    *,
    dataframe: pd.DataFrame,
    output_dir: Path,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    clean_df = dataframe.copy()

    for level in LEVEL_COLUMNS:
        clean_df[level] = clean_df[level].map(bool)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"INDIKATOR_BB_MB_BSH_BSB_EDITED_{timestamp}.xlsx"

    output_columns = [
        "Nama indikator",
        "Elemen",
        "Kategori",
        "No.",
        "Indikator",
        *LEVEL_COLUMNS,
    ]

    clean_df[output_columns].to_excel(output_path, index=False, sheet_name="Sheet1")

    return output_path


def build_name_mapping_dataframe(
    *,
    data_names: list[str],
    indicator_names: list[str],
) -> pd.DataFrame:
    rows = []
    indicator_keys = {normalize_key(name): name for name in indicator_names}
    indicator_options = list(indicator_names)

    for data_name in data_names:
        exact = indicator_keys.get(normalize_key(data_name), "")

        if not exact:
            match = get_close_matches(data_name, indicator_options, n=1, cutoff=0.65)
            exact = match[0] if match else ""

        rows.append(
            {
                "Nama data nilai": data_name,
                "Nama indikator": exact,
            }
        )

    return pd.DataFrame(rows)


def build_indicator_narratives_from_editor(
    *,
    indicator_df: pd.DataFrame,
    mapping_df: pd.DataFrame,
) -> dict[str, dict[str, str]]:
    result = {}

    if indicator_df.empty or mapping_df.empty:
        return result

    for _, mapping_row in mapping_df.iterrows():
        data_name = normalize_text(mapping_row.get("Nama data nilai"))
        indicator_name = normalize_text(mapping_row.get("Nama indikator"))

        if not data_name or not indicator_name:
            continue

        student_rows = indicator_df[
            indicator_df["Nama indikator"].map(normalize_key) == normalize_key(indicator_name)
        ]

        if student_rows.empty:
            continue

        nickname = nickname_from_name(data_name)
        narratives = {}

        for section_config in SECTION_CONFIGS_UI:
            section_name = section_config["name"]

            narratives[section_config["summary_header"]] = section_config["summary_template"].format(
                nickname=nickname
            )

            section_rows = student_rows[student_rows["Elemen"] == section_name]

            for category_letter in sorted(section_rows["Kategori"].dropna().unique()):
                category_rows = section_rows[section_rows["Kategori"] == category_letter]

                good_texts = category_rows[
                    category_rows["BSB"].fillna(False).astype(bool)
                ]["Indikator"].dropna().astype(str).tolist()

                bad_texts = category_rows[
                    category_rows["BB"].fillna(False).astype(bool)
                ]["Indikator"].dropna().astype(str).tolist()

                header = f"{section_config['category_prefix']} {category_letter}"

                narratives[header] = make_narrative(
                    good_texts=good_texts,
                    bad_texts=bad_texts,
                    guidance_word=section_config["guidance_word"],
                )

        result[normalize_key(data_name)] = narratives

    return result


# ==========================================================
# UI RENDER HELPERS
# ==========================================================

def render_indicator_editor_for_selected_student(
    *,
    selected_name: str,
    runtime_dir: Path,
) -> None:
    indicator_df = st.session_state["indicator_editor_df"]

    student_df = indicator_df[
        indicator_df["Nama indikator"].map(normalize_key) == normalize_key(selected_name)
    ].copy()

    if student_df.empty:
        st.warning("Data indikator untuk nama ini kosong.")
        return

    st.caption(
        "Aturan: satu baris hanya boleh memilih satu dari BB, MB, BSH, atau BSB. "
        "Kalau ingin mengganti pilihan, kosongkan pilihan pada baris itu dulu lalu Apply. "
        "Dalam satu nama + elemen + kategori, BB maksimal 1 dan BSB maksimal 1."
    )

    validation_errors = validate_indicator_editor(indicator_df)

    if validation_errors:
        st.warning("Ada kesalahan validasi. Perbaiki sebelum Submit.")
    else:
        st.success("Validasi indikator sementara aman.")

    for section_config in SECTION_CONFIGS_UI:
        section_name = section_config["name"]
        section_df = student_df[student_df["Elemen"] == section_name]

        if section_df.empty:
            continue

        st.markdown(f"### {section_name}")

        category_values = sorted(section_df["Kategori"].dropna().unique())

        for category_letter in category_values:
            category_df = section_df[section_df["Kategori"] == category_letter].copy()

            if category_df.empty:
                continue

            render_compact_category_editor(
                selected_name=selected_name,
                section_name=section_name,
                category_letter=str(category_letter),
                category_df=category_df,
            )

    action_col1, action_col2 = st.columns([1, 2])

    with action_col1:
        if st.button("Save Preview Indikator", use_container_width=True):
            try:
                output_path = save_indicator_editor_dataframe(
                    dataframe=st.session_state["indicator_editor_df"],
                    output_dir=runtime_dir,
                )

                st.session_state["edited_indicator_preview_path"] = str(output_path)
                st.success(f"Preview indikator berhasil disimpan: {output_path.name}")

            except Exception as exc:
                st.error(f"Simpan preview indikator gagal: {exc}")

    with action_col2:
        if "edited_indicator_preview_path" in st.session_state:
            output_path = Path(st.session_state["edited_indicator_preview_path"])

            if output_path.exists():
                st.download_button(
                    label="Download Preview Indikator Hasil Edit",
                    data=read_bytes(output_path),
                    file_name=output_path.name,
                    mime=EXCEL_MIME,
                    use_container_width=True,
                )

    with st.expander("Ringkasan validasi semua murid", expanded=False):
        validation_errors = validate_indicator_editor(st.session_state["indicator_editor_df"])

        if validation_errors:
            st.error("Masih ada kesalahan.")
            for error in validation_errors:
                st.write(f"- {error}")
        else:
            st.success("Semua indikator valid. Siap submit.")


# ==========================================================
# MAIN FEATURE 3
# ==========================================================

def render_feature_3_indicator_mapping(
    *,
    runtime_dir: Path,
    runtime_config: Any,
    max_students: int,
    target_semester_header: str,
) -> dict[str, Path | None]:
    st.subheader("3. Cocokkan data nilai dan indikator")

    st.write(
        "Bagian ini menggabungkan data nilai dengan indikator BB/MB/BSH/BSB. "
        "Narasi database tetap memakai BSB untuk bagian sangat baik dan BB untuk bagian bimbingan."
    )

    col1, col2, col3 = st.columns(3)

    with col1:
        student_file = st.file_uploader(
            "File data murid & nilai",
            type=["xlsx"],
            help="Opsional jika sudah memakai hasil edit dari bagian 2. Pakai Sheet1.",
        )

    with col2:
        database_template_file = st.file_uploader(
            "Template database semester 1",
            type=["xlsx"],
            help="Contoh: DATABASE SEMESTER 1.xlsx.",
        )

    with col3:
        checked_indicator_file = st.file_uploader(
            "File indikator/checklist",
            type=["xlsx"],
            help=(
                "Upload file indikator yang sudah berbentuk tabel horizontal. "
                "Bisa juga upload file preview indikator hasil edit."
            ),
        )

    saved_student_path: Path | None = None
    saved_database_template_path: Path | None = None
    saved_checked_indicator_path: Path | None = None

    if student_file:
        saved_student_path = save_uploaded_file(student_file, runtime_dir, "student_data.xlsx")

    if database_template_file:
        saved_database_template_path = save_uploaded_file(
            database_template_file,
            runtime_dir,
            "database_template.xlsx",
        )

    if checked_indicator_file:
        saved_checked_indicator_path = save_uploaded_file(
            checked_indicator_file,
            runtime_dir,
            "checked_indicator.xlsx",
        )

    if saved_student_path is None and "edited_student_file_path" in st.session_state:
        possible_edited_path = Path(st.session_state["edited_student_file_path"])

        if possible_edited_path.exists():
            saved_student_path = possible_edited_path

    indicator_path_for_preview: Path | None = None

    if saved_checked_indicator_path is not None:
        indicator_path_for_preview = saved_checked_indicator_path

    elif "edited_indicator_preview_path" in st.session_state:
        possible_indicator_path = Path(st.session_state["edited_indicator_preview_path"])

        if possible_indicator_path.exists():
            indicator_path_for_preview = possible_indicator_path

    elif "indicator_result_path" in st.session_state:
        possible_indicator_path = Path(st.session_state["indicator_result_path"])

        if possible_indicator_path.exists():
            indicator_path_for_preview = possible_indicator_path

    elif default_indicator_path().exists():
        indicator_path_for_preview = default_indicator_path()

    status_col1, status_col2, status_col3 = st.columns(3)

    with status_col1:
        if saved_student_path:
            st.success(f"Data nilai aktif: {saved_student_path.name}")
        else:
            st.warning("Data nilai belum ada. Upload file atau generate hasil edit di bagian 2.")

    with status_col2:
        if saved_database_template_path:
            st.success(f"Template database aktif: {saved_database_template_path.name}")
        else:
            st.warning("Template database belum diupload.")

    with status_col3:
        if indicator_path_for_preview:
            st.success(f"Indikator aktif: {indicator_path_for_preview.name}")
        else:
            st.warning("Indikator belum tersedia.")

    st.markdown("#### 3A. Preview dan edit indikator")

    can_load_indicator_preview = indicator_path_for_preview is not None

    if st.button(
        "See Preview Indikator",
        disabled=not can_load_indicator_preview,
        use_container_width=True,
    ):
        try:
            indicator_df = load_indicator_preview_dataframe(
                indicator_path=indicator_path_for_preview,
                config=runtime_config,
                target_semester_header=target_semester_header,
            )

            st.session_state["indicator_editor_df"] = indicator_df
            st.session_state["indicator_preview_loaded"] = True
            st.session_state["indicator_editor_version"] = 0

            if "name_mapping_df" in st.session_state:
                del st.session_state["name_mapping_df"]

            if "manual_indicator_narratives_by_name" in st.session_state:
                del st.session_state["manual_indicator_narratives_by_name"]

            st.success(
                f"Preview indikator berhasil dimuat. "
                f"Semester dibaca: {target_semester_header}. "
                "Tabel hanya muncul setelah tombol See Preview Indikator ditekan."
            )

        except Exception as exc:
            st.error(f"Load preview indikator gagal: {exc}")

    if st.session_state.get("indicator_preview_loaded") and "indicator_editor_df" in st.session_state:
        indicator_names = (
            st.session_state["indicator_editor_df"]["Nama indikator"]
            .dropna()
            .astype(str)
            .drop_duplicates()
            .tolist()
        )

        if indicator_names:
            selected_name = st.selectbox(
                "Pilih nama murid untuk preview indikator",
                options=indicator_names,
                key="selected_indicator_student_name",
            )

            render_indicator_editor_for_selected_student(
                selected_name=selected_name,
                runtime_dir=runtime_dir,
            )

    st.markdown("#### 3B. Cocokkan nama data nilai dengan nama indikator")

    if saved_student_path and st.session_state.get("indicator_preview_loaded") and "indicator_editor_df" in st.session_state:
        try:
            data_names = extract_student_names(
                workbook_path=saved_student_path,
                sheet_name=runtime_config.student_data_sheet,
                name_header=runtime_config.name_header,
                max_students=int(max_students),
            )

            indicator_names = (
                st.session_state["indicator_editor_df"]["Nama indikator"]
                .dropna()
                .astype(str)
                .drop_duplicates()
                .tolist()
            )

            if "name_mapping_df" not in st.session_state:
                st.session_state["name_mapping_df"] = build_name_mapping_dataframe(
                    data_names=data_names,
                    indicator_names=indicator_names,
                )

            mapping_df = st.data_editor(
                st.session_state["name_mapping_df"],
                key="name_mapping_editor",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Nama data nilai": st.column_config.Column(
                        "Nama data nilai",
                        pinned=True,
                        disabled=True,
                        width="medium",
                    ),
                    "Nama indikator": st.column_config.SelectboxColumn(
                        "Nama indikator",
                        options=["", *indicator_names],
                        width="medium",
                    ),
                },
                disabled=["Nama data nilai"],
            )

            st.session_state["name_mapping_df"] = mapping_df

        except Exception as exc:
            st.error(f"Preview mapping nama gagal: {exc}")

    elif not saved_student_path:
        st.info("Data nilai belum tersedia, jadi mapping nama belum bisa dibuat.")
    else:
        st.info("Klik See Preview Indikator dulu untuk membuat mapping nama.")

    st.markdown("#### 3C. Submit hasil indikator")

    can_submit_indicator = (
        st.session_state.get("indicator_preview_loaded")
        and "indicator_editor_df" in st.session_state
        and "name_mapping_df" in st.session_state
    )

    if st.button(
        "Submit Pilihan Indikator",
        disabled=not can_submit_indicator,
        type="primary",
        use_container_width=True,
    ):
        indicator_df = st.session_state["indicator_editor_df"]
        mapping_df = st.session_state["name_mapping_df"]

        validation_errors = validate_indicator_editor(indicator_df)

        unmapped_rows = mapping_df[
            mapping_df["Nama indikator"].isna()
            | (mapping_df["Nama indikator"].astype(str).str.strip() == "")
        ]

        if validation_errors:
            st.error("Submit gagal. Perbaiki centang indikator dulu.")
            for error in validation_errors:
                st.write(f"- {error}")

        elif not unmapped_rows.empty:
            st.error("Submit gagal. Masih ada nama data nilai yang belum dicocokkan dengan nama indikator.")
            st.dataframe(unmapped_rows, use_container_width=True, hide_index=True)

        else:
            manual_narratives = build_indicator_narratives_from_editor(
                indicator_df=indicator_df,
                mapping_df=mapping_df,
            )

            st.session_state["manual_indicator_narratives_by_name"] = manual_narratives
            st.success(
                "Pilihan indikator berhasil disubmit. "
                "Data ini akan dipakai saat Generate Database."
            )

    st.divider()

    return {
        "saved_student_path": saved_student_path,
        "saved_database_template_path": saved_database_template_path,
        "indicator_path_for_database": indicator_path_for_preview,
    }