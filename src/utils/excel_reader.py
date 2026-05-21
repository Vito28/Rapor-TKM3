from __future__ import annotations

from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_header(value: Any) -> str:
    return normalize(value).casefold()


def find_header_col(ws: Worksheet, header_name: str, scan_rows: int = 10) -> int:
    target = normalize_header(header_name)

    for row in range(1, min(ws.max_row, scan_rows) + 1):
        for col in range(1, ws.max_column + 1):
            if normalize_header(ws.cell(row=row, column=col).value) == target:
                return col

    # Fallback untuk variasi header seperti "Nama Murid" / "Nama Peserta".
    for row in range(1, min(ws.max_row, scan_rows) + 1):
        for col in range(1, ws.max_column + 1):
            value = normalize_header(ws.cell(row=row, column=col).value)
            if "nama" in value and ("peserta" in value or "murid" in value or "siswa" in value):
                return col

    raise ValueError(f"Kolom header '{header_name}' tidak ditemukan di sheet '{ws.title}'.")


def find_header_row(ws: Worksheet, required_header: str, scan_rows: int = 10) -> int:
    target = normalize_header(required_header)

    for row in range(1, min(ws.max_row, scan_rows) + 1):
        for col in range(1, ws.max_column + 1):
            if normalize_header(ws.cell(row=row, column=col).value) == target:
                return row

    for row in range(1, min(ws.max_row, scan_rows) + 1):
        for col in range(1, ws.max_column + 1):
            value = normalize_header(ws.cell(row=row, column=col).value)
            if "nama" in value and ("peserta" in value or "murid" in value or "siswa" in value):
                return row

    raise ValueError(f"Baris header '{required_header}' tidak ditemukan di sheet '{ws.title}'.")


def get_sheet(workbook_path: Path, sheet_name: str | None = None) -> Worksheet:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.active


def extract_student_names(
    workbook_path: Path,
    sheet_name: str = "Sheet1",
    name_header: str = "Nama peserta didik",
    max_students: int | None = None,
) -> list[str]:
    ws = get_sheet(workbook_path, sheet_name)
    header_row = find_header_row(ws, name_header)
    name_col = find_header_col(ws, name_header)

    names: list[str] = []
    for row in range(header_row + 1, ws.max_row + 1):
        name = normalize(ws.cell(row=row, column=name_col).value)
        if name:
            names.append(name)
        if max_students and len(names) >= max_students:
            break

    return names


def read_rows_by_header(
    workbook_path: Path,
    sheet_name: str = "Sheet1",
    required_header: str = "Nama peserta didik",
    max_rows: int | None = None,
) -> list[dict[str, Any]]:
    ws = get_sheet(workbook_path, sheet_name)
    header_row = find_header_row(ws, required_header)

    headers: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        header = normalize(ws.cell(row=header_row, column=col).value)
        if header:
            headers[col] = header

    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data: dict[str, Any] = {}
        has_value = False

        for col_idx, header in headers.items():
            value = ws.cell(row=row_idx, column=col_idx).value
            if value not in (None, ""):
                has_value = True
            row_data[header] = value

        if has_value:
            rows.append(row_data)

        if max_rows and len(rows) >= max_rows:
            break

    return rows


def nickname_from_name(name: str) -> str:
    name = normalize(name)
    if not name:
        return ""
    return name.split()[0]
