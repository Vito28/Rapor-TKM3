from __future__ import annotations

from copy import copy
from pathlib import Path
import re

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

# ==========================================================
# KONFIGURASI FILE
# ==========================================================
INPUT_RUBRIC = Path("data/input/RAPOR HIJAU_INDIKATOR.xlsx")
INPUT_DATABASE_TEMPLATE = Path("data/input/DATABASE SEMESTER 1.xlsx")
OUTPUT_DATABASE = Path("data/output/DATABASE_SEMESTER_1_GENERATED.xlsx")

SOURCE_SHEET_NAME = "Sheet1"
DATABASE_SHEET_NAME = "Sheet 1"

# Sheet1: nama anak diisi di baris tepat di atas header SEMESTER I/II.
# Pada file kamu: header SEMESTER ada di row 4, jadi nama anak di row 3.
NAME_ROW_OFFSET_FROM_SEMESTER_ROW = -1
SCORE_ROW_OFFSET_FROM_SEMESTER_ROW = 1

# Hanya Semester I yang diproses.
TARGET_SEMESTER_HEADER = "SEMESTER I"
TARGET_SEMESTER_VALUE = "1 (satu)"

# Kolom teks indikator/capaian pada Sheet1.
CAPABILITY_TEXT_COLUMN = 2  # B

# Nilai yang dipakai untuk narasi.
GOOD_LEVEL = "BSB"
BAD_LEVEL = "BB"

# Kalau True, blok Semester I yang nama anaknya kosong akan dilewati.
# Kalau False, script akan membuat nama placeholder, misalnya BELUM_DIISI_F.
SKIP_EMPTY_NAME_BLOCKS = False

CHECK_SYMBOLS = {"✓", "✔", "☑", "√", "ü"}

SECTION_CONFIGS = [
    {
        "marker": "I. ELEMEN NILAI AGAMA DAN BUDI PEKERTI",
        "name": "Agama",
        "summary_header": "Agama ",
        "category_prefix": "Agama",
        "summary_template": "Pada umumnya nilai agama dan moral {nickname} sudah baik.",
        "guidance_word": "nasihat",
    },
    {
        "marker": "II. ELEMEN JATI DIRI",
        "name": "Jati Diri",
        "summary_header": "Jati Diri ",
        "category_prefix": "Jati Diri",
        "summary_template": "Pada umumnya kemampuan jati diri {nickname} sudah baik.",
        "guidance_word": "latihan",
    },
    {
        "marker": "III. ELEMEN DASAR-DASAR LITERASI, MATEMATIKA, SAINS, TEKNOLOGI, REKAYASA DAN SENI",
        "name": "STEAM",
        "summary_header": "STEAM ",
        "category_prefix": "STEAM",
        "summary_template": "Pada umumnya kemampuan dasar-dasar literasi, matematika, sains, teknologi, rekayasa dan seni {nickname} sudah baik.",
        "guidance_word": "latihan",
    },
]


# ==========================================================
# HELPER CELL / TEXT
# ==========================================================

def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_header(value) -> str:
    return normalize_text(value).casefold()


def get_merged_value(ws: Worksheet, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value

    return None


def get_cell_text(ws: Worksheet, row: int, col: int) -> str:
    return normalize_text(get_merged_value(ws, row, col))


def is_letter_label(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Z]\.?", normalize_text(value).upper()))


def is_number_label(value: str) -> bool:
    return bool(re.fullmatch(r"\d+\.", normalize_text(value)))


def clean_capability_text(text: str) -> str:
    text = normalize_text(text).rstrip(".")
    text = re.sub(r"^\d+\.\s*", "", text)
    if text:
        text = text[0].lower() + text[1:]
    return text


def join_phrases(items: list[str]) -> str:
    cleaned: list[str] = []
    for item in items:
        text = clean_capability_text(item)
        if text:
            cleaned.append(text)

    if not cleaned:
        return ""
    if len(cleaned) == 1:
        return cleaned[0]
    if len(cleaned) == 2:
        return f"{cleaned[0]} dan {cleaned[1]}"
    return f"{', '.join(cleaned[:-1])}, dan {cleaned[-1]}"


def nickname_from_name(name: str) -> str:
    name = normalize_text(name)
    if not name or name.startswith("BELUM_DIISI"):
        return "anak"
    return name.split()[0]


def is_checkmark(formula_value, real_value) -> bool:
    values = []
    if formula_value is not None:
        values.append(str(formula_value).strip())
    if real_value is not None:
        values.append(str(real_value).strip())

    for value in values:
        raw = value.strip()
        upper = raw.upper().replace(" ", "")

        if raw in CHECK_SYMBOLS:
            return True
        if "UNICHAR(10003)" in upper:
            return True
        if "UNICHAR(10004)" in upper:
            return True
        if "CHAR(252)" in upper:
            return True

    return False


# ==========================================================
# DETEKSI STRUKTUR SHEET1
# ==========================================================

def find_semester_row(ws: Worksheet) -> int:
    for row in range(1, min(ws.max_row, 40) + 1):
        for col in range(1, ws.max_column + 1):
            if "SEMESTER" in get_cell_text(ws, row, col).upper():
                return row
    raise ValueError("Baris header SEMESTER tidak ditemukan di Sheet1.")


def build_semester_block(ws: Worksheet, semester_row: int, start_col: int, end_col: int) -> dict | None:
    semester_name = get_cell_text(ws, semester_row, start_col).upper()
    if semester_name != TARGET_SEMESTER_HEADER:
        return None

    score_row = semester_row + SCORE_ROW_OFFSET_FROM_SEMESTER_ROW
    name_row = semester_row + NAME_ROW_OFFSET_FROM_SEMESTER_ROW

    score_columns: dict[str, int] = {}
    for col in range(start_col, end_col + 1):
        level = get_cell_text(ws, score_row, col).upper()
        if level in {"BB", "MB", "BSH", "BSB"}:
            score_columns[level] = col

    if GOOD_LEVEL not in score_columns or BAD_LEVEL not in score_columns:
        return None

    name = ""
    for col in range(start_col, end_col + 1):
        name = get_cell_text(ws, name_row, col)
        if name:
            break

    if not name:
        if SKIP_EMPTY_NAME_BLOCKS:
            return None
        name = f"BELUM_DIISI_{get_column_letter(start_col)}"

    return {
        "name": name,
        "semester": TARGET_SEMESTER_VALUE,
        "start_col": start_col,
        "end_col": end_col,
        "score_columns": score_columns,
    }


def find_semester_1_blocks(ws: Worksheet) -> list[dict]:
    semester_row = find_semester_row(ws)
    blocks: list[dict] = []
    used_ranges: set[tuple[int, int]] = set()

    # Utama: baca dari merged cell header SEMESTER I.
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row != semester_row:
            continue

        value = get_cell_text(ws, merged_range.min_row, merged_range.min_col).upper()
        if value != TARGET_SEMESTER_HEADER:
            continue

        start_col, end_col = merged_range.min_col, merged_range.max_col
        used_ranges.add((start_col, end_col))

        block = build_semester_block(ws, semester_row, start_col, end_col)
        if block:
            blocks.append(block)

    # Fallback: kalau header SEMESTER I tidak di-merge.
    for col in range(1, ws.max_column + 1):
        value = get_cell_text(ws, semester_row, col).upper()
        if value != TARGET_SEMESTER_HEADER:
            continue

        if any(start <= col <= end for start, end in used_ranges):
            continue

        block = build_semester_block(ws, semester_row, col, col + 3)
        if block:
            blocks.append(block)

    blocks.sort(key=lambda block: block["start_col"])
    return blocks


def find_section_rows(ws: Worksheet) -> dict[str, tuple[int, int]]:
    """Return mapping section_name -> (start_row, end_row)."""
    markers: list[tuple[int, dict]] = []

    for row in range(1, ws.max_row + 1):
        label = get_cell_text(ws, row, 1).upper()
        for config in SECTION_CONFIGS:
            if label.startswith(config["marker"].upper()):
                markers.append((row, config))
                break

    if not markers:
        raise ValueError("Marker elemen I/II/III tidak ditemukan di Sheet1.")

    result: dict[str, tuple[int, int]] = {}
    for idx, (start_row, config) in enumerate(markers):
        end_row = markers[idx + 1][0] - 1 if idx + 1 < len(markers) else ws.max_row
        result[config["name"]] = (start_row, end_row)

    return result


def detect_categories(ws: Worksheet, section_start: int, section_end: int) -> dict[str, list[int]]:
    """
    Deteksi kategori A/B/C/D/G dalam sebuah elemen.
    Isi kategori hanya baris indikator bernomor: 1., 2., 3., dst.
    """
    categories: dict[str, list[int]] = {}
    current_letter = ""

    for row in range(section_start + 1, section_end + 1):
        label = get_cell_text(ws, row, 1).upper().replace(" ", "")
        text = get_cell_text(ws, row, CAPABILITY_TEXT_COLUMN)

        if is_letter_label(label):
            current_letter = label.replace(".", "")
            categories[current_letter] = []
            continue

        if current_letter and is_number_label(label) and text:
            categories[current_letter].append(row)

    return categories


# ==========================================================
# EKSTRAK NARASI
# ==========================================================

def extract_checked_texts(
    ws_formula: Worksheet,
    ws_value: Worksheet,
    indicator_rows: list[int],
    block: dict,
) -> tuple[list[str], list[str]]:
    good_texts: list[str] = []
    bad_texts: list[str] = []

    good_col = block["score_columns"][GOOD_LEVEL]
    bad_col = block["score_columns"][BAD_LEVEL]

    for row in indicator_rows:
        capability_text = get_cell_text(ws_formula, row, CAPABILITY_TEXT_COLUMN)
        if not capability_text:
            continue

        good_formula = get_merged_value(ws_formula, row, good_col)
        good_real = get_merged_value(ws_value, row, good_col)
        bad_formula = get_merged_value(ws_formula, row, bad_col)
        bad_real = get_merged_value(ws_value, row, bad_col)

        if is_checkmark(good_formula, good_real):
            good_texts.append(capability_text)
        if is_checkmark(bad_formula, bad_real):
            bad_texts.append(capability_text)

    return good_texts, bad_texts


def make_narrative(good_texts: list[str], bad_texts: list[str], guidance_word: str) -> str:
    good_part = join_phrases(good_texts)
    bad_part = join_phrases(bad_texts)

    if good_part and bad_part:
        return (
            f"Dalam hal {good_part} sangat baik, "
            f"namun dalam hal {bad_part} masih perlu bimbingan dan {guidance_word}."
        )
    if good_part:
        return f"Dalam hal {good_part} sudah sangat baik."
    if bad_part:
        return f"Dalam hal {bad_part} masih perlu bimbingan dan {guidance_word}."
    return ""


def generate_narratives_for_block(ws_formula: Worksheet, ws_value: Worksheet, block: dict) -> dict[str, str]:
    section_ranges = find_section_rows(ws_formula)
    output: dict[str, str] = {}
    nickname = nickname_from_name(block["name"])

    for config in SECTION_CONFIGS:
        section_name = config["name"]
        start_row, end_row = section_ranges[section_name]
        categories = detect_categories(ws_formula, start_row, end_row)

        output[config["summary_header"]] = config["summary_template"].format(nickname=nickname)

        for letter, indicator_rows in categories.items():
            good_texts, bad_texts = extract_checked_texts(
                ws_formula=ws_formula,
                ws_value=ws_value,
                indicator_rows=indicator_rows,
                block=block,
            )
            header = f"{config['category_prefix']} {letter}"
            output[header] = make_narrative(good_texts, bad_texts, config["guidance_word"])

    return output


# ==========================================================
# DATABASE OUTPUT
# ==========================================================

def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def load_or_create_database_workbook() -> tuple[Workbook, Worksheet]:
    if INPUT_DATABASE_TEMPLATE.exists():
        wb = load_workbook(INPUT_DATABASE_TEMPLATE)
        ws = wb[DATABASE_SHEET_NAME] if DATABASE_SHEET_NAME in wb.sheetnames else wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = DATABASE_SHEET_NAME
    raise FileNotFoundError(
        f"Template database tidak ditemukan: {INPUT_DATABASE_TEMPLATE}. "
        "Simpan DATABASE SEMESTER 1.xlsx di folder data/input."
    )


def header_map(ws: Worksheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_text(ws.cell(1, col).value)
        if header:
            mapping[normalize_header(header)] = col
    return mapping


def get_header_col(headers: dict[str, int], header_name: str) -> int | None:
    return headers.get(normalize_header(header_name))


def existing_rows_by_name(ws: Worksheet, headers: dict[str, int]) -> dict[str, int]:
    name_col = get_header_col(headers, "Nama peserta didik")
    if not name_col:
        return {}

    result = {}
    for row in range(2, ws.max_row + 1):
        name = normalize_text(ws.cell(row, name_col).value)
        if name:
            result[normalize_header(name)] = row
    return result


def clear_data_rows(ws: Worksheet) -> None:
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)


def write_database_row(ws: Worksheet, headers: dict[str, int], row_index: int, row_data: dict[str, object]) -> None:
    for header, value in row_data.items():
        col = get_header_col(headers, header)
        if col:
            ws.cell(row_index, col).value = value


def build_output() -> None:
    if not INPUT_RUBRIC.exists():
        raise FileNotFoundError(f"File rubrik tidak ditemukan: {INPUT_RUBRIC}")

    wb_formula = load_workbook(INPUT_RUBRIC, data_only=False)
    wb_value = load_workbook(INPUT_RUBRIC, data_only=True)

    if SOURCE_SHEET_NAME not in wb_formula.sheetnames:
        raise ValueError(f"Sheet '{SOURCE_SHEET_NAME}' tidak ditemukan. Sheet tersedia: {wb_formula.sheetnames}")

    ws_formula = wb_formula[SOURCE_SHEET_NAME]
    ws_value = wb_value[SOURCE_SHEET_NAME]

    blocks = find_semester_1_blocks(ws_formula)
    if not blocks:
        raise ValueError("Tidak ada blok SEMESTER I yang valid di Sheet1.")

    db_wb, db_ws = load_or_create_database_workbook()
    headers = header_map(db_ws)

    # Simpan data lama berdasarkan nama, agar biodata lama bisa ikut kalau namanya cocok.
    old_rows_by_name = existing_rows_by_name(db_ws, headers)
    old_data: dict[str, list[object]] = {}
    for norm_name, row_idx in old_rows_by_name.items():
        old_data[norm_name] = [db_ws.cell(row_idx, col).value for col in range(1, db_ws.max_column + 1)]

    style_source_row = 2 if db_ws.max_row >= 2 else 1
    max_col = db_ws.max_column
    clear_data_rows(db_ws)

    for output_idx, block in enumerate(blocks, start=2):
        name = block["name"]
        norm_name = normalize_header(name)

        if output_idx != style_source_row:
            copy_row_style(db_ws, style_source_row if style_source_row <= db_ws.max_row else 1, output_idx, max_col)

        # Kalau nama cocok dengan database lama, clone satu baris biodata lama dulu.
        if norm_name in old_data:
            for col, value in enumerate(old_data[norm_name], start=1):
                db_ws.cell(output_idx, col).value = value

        narratives = generate_narratives_for_block(ws_formula, ws_value, block)

        row_data: dict[str, object] = {
            "No.": f"{output_idx - 1}.",
            "Nama peserta didik": name,
            "Nama panggilan": nickname_from_name(name),
            "Semester": TARGET_SEMESTER_VALUE,
            **narratives,
        }

        write_database_row(db_ws, headers, output_idx, row_data)

    OUTPUT_DATABASE.parent.mkdir(parents=True, exist_ok=True)
    db_wb.save(OUTPUT_DATABASE)

    print("Selesai.")
    print(f"Sumber rubrik       : {INPUT_RUBRIC}")
    print(f"Sheet diproses      : {SOURCE_SHEET_NAME}")
    print(f"Semester diproses   : {TARGET_SEMESTER_HEADER}")
    print(f"Jumlah murid/blok   : {len(blocks)}")
    print(f"Output database     : {OUTPUT_DATABASE}")


if __name__ == "__main__":
    build_output()
