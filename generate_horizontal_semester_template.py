from __future__ import annotations

from copy import copy
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ==========================================================
# KONFIGURASI
# ==========================================================
INPUT_FILE = Path("data/input/RAPOR HIJAU_INDIKATOR.xlsx")
OUTPUT_FILE = Path("data/output/FORMAT_INDIKATOR_21_MURID.xlsx")

TEMPLATE_SHEET = "Sheet1"      # sheet indikator/rubrik asli
STUDENT_SHEET = "Sheet3"       # sheet daftar 21 murid
NAME_HEADER = "Nama peserta didik"

MAX_STUDENTS = 21
LEFT_COPY_UNTIL_COL = 5         # A:E = bagian NO + CAPAIAN PEMBELAJARAN
FIRST_STUDENT_COL = 6           # F
COLUMNS_PER_STUDENT = 8         # Semester I 4 kolom + Semester II 4 kolom

NAME_ROW = 3                    # baris nama anak
SEMESTER_ROW = 4                # baris SEMESTER I / SEMESTER II
SCORE_ROW = 5                   # baris BB MB BSH BSB
DATA_START_ROW = 6              # mulai isi centang

SEMESTER_LABELS = ["SEMESTER I", "SEMESTER II"]
SCORE_LABELS = ["BB", "MB", "BSH", "BSB"]


def normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def find_header_col(ws, header_name: str, scan_rows: int = 10) -> int:
    target = normalize(header_name).lower()

    for row in range(1, min(ws.max_row, scan_rows) + 1):
        for col in range(1, ws.max_column + 1):
            value = normalize(ws.cell(row=row, column=col).value).lower()
            if value == target:
                return col

    # fallback: cari header yang mengandung nama
    for row in range(1, min(ws.max_row, scan_rows) + 1):
        for col in range(1, ws.max_column + 1):
            value = normalize(ws.cell(row=row, column=col).value).lower()
            if "nama" in value and "peserta" in value:
                return col

    raise ValueError(f"Kolom header '{header_name}' tidak ditemukan di sheet {ws.title}.")


def get_student_names(ws) -> list[str]:
    name_col = find_header_col(ws, NAME_HEADER)
    names: list[str] = []

    for row in range(2, ws.max_row + 1):
        name = normalize(ws.cell(row=row, column=name_col).value)
        if name:
            names.append(name)

        if len(names) >= MAX_STUDENTS:
            break

    if not names:
        raise ValueError(f"Tidak ada nama murid di sheet {ws.title}.")

    return names


def copy_cell_style(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)

    if src_cell.hyperlink:
        dst_cell._hyperlink = copy(src_cell.hyperlink)

    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)


def copy_left_rubric(src_ws, dst_ws) -> None:
    """Copy bagian kiri A:E dari Sheet1 asli ke file output baru."""
    for row in range(1, src_ws.max_row + 1):
        for col in range(1, LEFT_COPY_UNTIL_COL + 1):
            src_cell = src_ws.cell(row=row, column=col)
            dst_cell = dst_ws.cell(row=row, column=col)
            dst_cell.value = src_cell.value
            copy_cell_style(src_cell, dst_cell)

    # Copy lebar kolom A:E
    for col in range(1, LEFT_COPY_UNTIL_COL + 1):
        col_letter = get_column_letter(col)
        dst_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width

    # Copy tinggi baris
    for row in range(1, src_ws.max_row + 1):
        dst_ws.row_dimensions[row].height = src_ws.row_dimensions[row].height

    # Copy merged cells yang berada sepenuhnya di A:E
    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.max_col <= LEFT_COPY_UNTIL_COL:
            dst_ws.merge_cells(str(merged_range))


def style_generated_area(ws, max_row: int, first_col: int, last_col: int) -> None:
    thin = Side(style="thin", color="000000")
    dotted = Side(style="dotted", color="808080")

    for row in range(1, max_row + 1):
        for col in range(first_col, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if row >= DATA_START_ROW:
                cell.border = Border(left=dotted, right=dotted, top=dotted, bottom=dotted)
            else:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header nama dan semester
    dark_fill = PatternFill("solid", fgColor="D9EAF7")
    score_fill = PatternFill("solid", fgColor="F2F2F2")

    for col in range(first_col, last_col + 1):
        ws.cell(NAME_ROW, col).font = Font(bold=True)
        ws.cell(NAME_ROW, col).fill = dark_fill
        ws.cell(SEMESTER_ROW, col).font = Font(bold=True)
        ws.cell(SEMESTER_ROW, col).fill = dark_fill
        ws.cell(SCORE_ROW, col).font = Font(bold=True, size=8)
        ws.cell(SCORE_ROW, col).fill = score_fill

    # Lebar kolom nilai dibuat ramping
    for col in range(first_col, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 5


def generate_student_blocks(ws, names: list[str], max_row: int) -> None:
    total_generated_cols = len(names) * COLUMNS_PER_STUDENT
    last_col = FIRST_STUDENT_COL + total_generated_cols - 1

    style_generated_area(ws, max_row, FIRST_STUDENT_COL, last_col)

    for student_index, name in enumerate(names):
        start_col = FIRST_STUDENT_COL + student_index * COLUMNS_PER_STUDENT
        end_col = start_col + COLUMNS_PER_STUDENT - 1

        # Nama murid, merge di atas SEMESTER I + SEMESTER II
        safe_merge_cells(
            ws,
            start_row=NAME_ROW,
            start_column=start_col,
            end_row=NAME_ROW,
            end_column=end_col,
        )
        ws.cell(NAME_ROW, start_col).value = name

        # SEMESTER I dan SEMESTER II
        sem1_start = start_col
        sem1_end = start_col + 3
        sem2_start = start_col + 4
        sem2_end = start_col + 7

        safe_merge_cells(
            ws,
            start_row=SEMESTER_ROW,
            start_column=sem1_start,
            end_row=SEMESTER_ROW,
            end_column=sem1_end,
        )
        ws.cell(SEMESTER_ROW, sem1_start).value = SEMESTER_LABELS[0]

        safe_merge_cells(
            ws,
            start_row=SEMESTER_ROW,
            start_column=sem2_start,
            end_row=SEMESTER_ROW,
            end_column=sem2_end,
        )
        ws.cell(SEMESTER_ROW, sem2_start).value = SEMESTER_LABELS[1]

        # BB MB BSH BSB untuk semester I dan II
        for offset, label in enumerate(SCORE_LABELS):
            ws.cell(SCORE_ROW, sem1_start + offset).value = label
            ws.cell(SCORE_ROW, sem2_start + offset).value = label

    ws.freeze_panes = "F6"

def safe_merge_cells(ws, start_row, start_column, end_row, end_column):
    try:
        ws.merge_cells(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column,
        )
    except IndexError:
        # Bersihkan border area merge, lalu coba merge ulang
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                ws.cell(row=row, column=col).border = Border()

        ws.merge_cells(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column,
        )


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"File input tidak ditemukan: {INPUT_FILE}")

    source_wb = load_workbook(INPUT_FILE)

    if TEMPLATE_SHEET not in source_wb.sheetnames:
        raise ValueError(f"Sheet template '{TEMPLATE_SHEET}' tidak ditemukan. Sheet tersedia: {source_wb.sheetnames}")

    if STUDENT_SHEET not in source_wb.sheetnames:
        raise ValueError(
            f"Sheet data murid '{STUDENT_SHEET}' tidak ditemukan. "
            f"Buat sheet bernama '{STUDENT_SHEET}' dengan kolom '{NAME_HEADER}'. "
            f"Sheet tersedia: {source_wb.sheetnames}"
        )

    src_ws = source_wb[TEMPLATE_SHEET]
    student_ws = source_wb[STUDENT_SHEET]

    names = get_student_names(student_ws)

    output_wb = Workbook()
    out_ws = output_wb.active
    out_ws.title = "Sheet1"

    copy_left_rubric(src_ws, out_ws)
    generate_student_blocks(out_ws, names, src_ws.max_row)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    output_wb.save(OUTPUT_FILE)

    print("Selesai.")
    print(f"Jumlah murid dibuat: {len(names)}")
    print(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
